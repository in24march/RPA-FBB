import logging
import math
import pandas as pd
import time
import openpyxl
from openpyxl import load_workbook, Workbook
from copy import copy
from datetime import datetime, timedelta
from openpyxl.styles import Alignment
from setting_fbb import *
import setting_fbb

def Play_premium(promotion, task_type, date):
    finder_file = find_file(PP_path)
    finder_data = find_file(Data_path)
    file_PP = finder_file.file_last_time()
    file_data = finder_data.file_last_time()
    wb = load_workbook(file_PP)
    ws = wb.active
    df = pd.read_excel(file_data)
    
    filtered_data = df[df['Contact Status'] == 'Success Closure']
    number_data = filtered_data['Contact Number']
    
    #แก้เบอ เพิ่ม0 
    if task_type == 'Mobile':
        formatted_numbers = number_data.astype(str).apply(lambda x: x.zfill(10)[:10])
    else:  # task_type == 'FBB'
        formatted_numbers = number_data.astype(str)
    
    start_row = 2
    max_row = ws.max_row
    print(formatted_numbers)
    
    #เอาเบอร์ที่ filtered แล้วมาใส่ใน Play Premium
    for i, number in enumerate(formatted_numbers, start= start_row):
        ws[f'D{i}'] = number
    
    for j in range(i + 1, max_row + 1):
        ws[f'D{j}'] = None
    
    #ทำให้ row เท่ากันกันข้อมูล
    column_to_adjust = ['A', 'B', 'C', 'E', 'F', 'G', 'AO', 'GL']
    for col in column_to_adjust:
        for j in range(start_row, max_row + 1):
            if j > i:
                ws[f'{col}{j}'] = None
            elif j > i - len(formatted_numbers):
                ws[f'{col}{j}'] = ws[f'{col}{start_row}'].value
                ws[f'{col}{j}'].font = copy(ws[f'{col}{start_row}'].font)
    
    date_format = date.strftime("%d/%m/%Y")
    for k in range(start_row, start_row + len(formatted_numbers)):
        ws[f'F{k}'] = promotion
        ws[f'G{k}'] = date_format
        
    file_date_name = date.strftime("%d%m%Y")
    new_file_name = f"File Play Premium Plus {task_type} {file_date_name}.xlsx"
    
    # เซฟไฟล์เป็นชื่อใหม่
    new_file_path = os.path.join(sub3_folder, new_file_name)
    wb.save(new_file_path)
    
    
    
def patch_data():
    finder_file = find_file(Patch_data)
    PD_file = finder_file.file_last_time()
    wb = load_workbook(PD_file)
    ws = wb['Use Cases']
    df = pd.read_excel(PD_file, sheet_name='Use Cases')
    
    # for row in ws.iter_rows(min_row= 2, max_row= ws.max_row, min_col= 1 ,max_col= ws.max_column):
    #     call_attempt = row[4].value
    #     contact_status = row[2].value
    #     call_outcome = row[3].value
        
    #     if call_attempt == 5 and (contact_status == 'Unsuccess' or not contact_status):
    #         row[2].value = 'Success SMS'
        
    #     if row[2].value == 'Success SMS':
    #         row[3].value = None
        
    #     if not contact_status:
    #         row[2].value = 'Unsuccess'
        
    #     row[4].alignment = Alignment(horizontal='left')
        
    # wb.save(PD_file)
    
    df.loc[(df['Call Attempt'] == 5) & ((df['Contact Status'] == 'Unsuccess') | (df['Contact Status'].isnull())), 'Contact Status'] = 'Success SMS'
    df.loc[df['Contact Status'] == 'Success SMS', 'Call Outcome'] = None
    df.loc[df['Contact Status'].isnull(), 'Contact Status'] = 'Unsuccess'

    df.to_excel(PD_file, sheet_name='Use Cases', index=False)
    df_number = pd.read_excel(PD_file, sheet_name= 'Use Cases')
    success_closure(df_number, PD_file)
    
def success_closure(df_number, file_path):
    # filtered_number = df_number[df_number['Contact Status'].isin(['Success', 'Success SMS'])]
    # closure_num = df_number[df_number['Contact Status'] == 'Success Closure']
    # tel_num = filtered_number['Mobile number']
    # tel_closure = closure_num['Mobile number']
    # print(tel_num.head(5))
    
    # wb = load_workbook(file_path)
    # if 'Success & Success SMS' not in wb.sheetnames:
    #     ws_success = wb.create_sheet('Success & Success SMS')
    # else:
    #     ws_success = wb['Success & Success SMS']
    
    # # Write tel_num to the sheet
    # for index, num in enumerate(tel_num, start=1):
    #     ws_success.cell(row=index, column=1, value=num)
    
    # if 'Closure' not in wb.sheetnames:
    #     ws_closure = wb.create_sheet('Closure')
    # else:
    #     ws_closure = wb['Closure']
    
    # # Write tel_num to the 'Closure' sheet
    # for index, num in enumerate(tel_closure, start=1):
    #     ws_closure.cell(row=index, column=1, value=num)
        
    # wb.save(file_path)
    filtered_number = df_number[df_number['Contact Status'].isin(['Success', 'Success SMS'])]
    closure_num = df_number[df_number['Contact Status'] == 'Success Closure']
    tel_num = filtered_number['Mobile number']
    tel_closure = closure_num['Mobile number']
    
    print(tel_num.head(5))  # แสดงตัวอย่างข้อมูลที่ถูกกรอง

    # โหลดไฟล์ Excel ด้วย openpyxl
    wb = load_workbook(file_path)
    
    # เขียนข้อมูลลงใน sheet "Success & Success SMS"
    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        filtered_number[['Mobile number']].to_excel(writer, sheet_name='Success & Success SMS', index=False, header=False)

    # เขียนข้อมูลลงใน sheet "Closure"
    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        closure_num[['Mobile number']].to_excel(writer, sheet_name='Closure', index=False, header=False)
    
def change_outcome(word):
    # finder_file = find_file(Patch_data)
    # PD_file = finder_file.file_last_time()
    # wb = load_workbook(PD_file)
    # ws = wb['Use Cases']
    # for row in ws.iter_rows(min_row= 2, max_row= ws.max_row, min_col= 1 ,max_col= ws.max_column):
    #     contact_status = row[2].value
    #     call_outcome = row[3].value
        
    #     if contact_status == 'Unsuccess':
    #         row[3].value = word
            
    # print(f'change outcome')
    # wb.save(PD_file)
    finder_file = find_file(Patch_data)
    PD_file = finder_file.file_last_time()

    # อ่านข้อมูลทั้งหมดลงใน pandas DataFrame
    df = pd.read_excel(PD_file, sheet_name='Use Cases')

    # เปลี่ยนค่า call_outcome ตามเงื่อนไข
    df.loc[df['Contact Status'] == 'Unsuccess', 'Call Outcome'] = word

    # เขียนข้อมูลกลับเข้าไปใน Excel
    with pd.ExcelWriter(PD_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df.to_excel(writer, sheet_name='Use Cases', index=False)
    
    print('change outcome')
    
def split_data():
    finder_file = find_file(Patch_data)
    PD_file = finder_file.file_last_time()
    df_number = pd.read_excel(PD_file, sheet_name= 'Use Cases')
    wb = load_workbook(PD_file)
    ws = wb['Use Cases']
    
    # total_rows = ws.max_row
    # max_row_per_file = 100000
    # num_file = math.ceil(total_rows / max_row_per_file)
    
    # for i in range(num_file):
    #     start_rows = i * max_row_per_file + 1
    #     end_rows = min((i + 1) * max_row_per_file, total_rows)
        
    #     wb_new = Workbook()
    #     ws_new = wb_new.active
        
    #     header = [cell.value for cell in ws[1]]
    #     ws_new.append(header)
        
    #     if i == 0:
    #         for row_idx, row in enumerate(ws.iter_rows(min_row=start_rows + 1, max_row=end_rows, values_only=True), start=2):
    #             ws_new.append(row)
    #             ws_new.cell(row=row_idx, column=5).alignment = Alignment(horizontal='left')
    #     else:
    #         for row_idx, row in enumerate(ws.iter_rows(min_row=start_rows, max_row=end_rows, values_only=True), start=2):
    #             ws_new.append(row)
    #             ws_new.cell(row=row_idx, column=5).alignment = Alignment(horizontal='left')
    #     # สร้างชื่อไฟล์ใหม่ใน P_data
    #     new_file_name = os.path.join(P_data, os.path.basename(PD_file).replace(".xlsx", f"_{i+1}.xlsx"))
    #     # บันทึกไฟล์ในโฟลเดอร์ P_data
    #     wb_new.save(new_file_name)
        
    #     df_number = pd.read_excel(new_file_name, sheet_name= 'Sheet')
    #     success_closure(df_number, new_file_name)
        
    #     print(f"Save file: {new_file_name}")
        
    #     df_new = pd.read_excel(new_file_name, sheet_name= 'Sheet')
        
    #     txt_file_name = os.path.join(txt_path, os.path.basename(new_file_name).replace(".xlsx", ".txt"))
        
    #     df_new_filtered = df_new.drop(columns=['Mobile number'], errors='ignore')
        
    #     with open(txt_file_name, 'w') as file:
            
    #         header_line = '|'.join(df_new_filtered.columns)
    #         file.write(header_line + '\n')
            
    #         df_new_filtered = df_new_filtered.fillna('')
            
    #         for index, row in df_new_filtered.iterrows():
    #             line = '|'.join(map(str, row.values))
    #             file.write(line + '\n')
                
    #     print(f"Save file txt: {txt_file_name}")
    df_number = pd.read_excel(PD_file, sheet_name='Use Cases')
    total_rows = len(df_number)
    
    # กำหนดจำนวนแถวสูงสุดต่อไฟล์
    max_row_per_file = 100000
    num_file = math.ceil(total_rows / max_row_per_file)
    
    for i in range(num_file):
        # หาขอบเขตแถวที่จะใส่ในแต่ละไฟล์
        start_rows = i * max_row_per_file
        end_rows = min((i + 1) * max_row_per_file, total_rows)
        
        # กรองข้อมูลเฉพาะแถวในขอบเขต
        df_chunk = df_number.iloc[start_rows:end_rows]
        
        # สร้างไฟล์ Excel ใหม่
        new_file_name = os.path.join(sub1_folder, os.path.basename(PD_file).replace(".xlsx", f"_{i+1}.xlsx")) #เปลี่ยนตรงนี้ xlsx patch_data
        df_chunk.to_excel(new_file_name, index=False)
        
        # ใช้ success_closure กับข้อมูลใหม่
        success_closure(df_chunk, new_file_name)
        
        print(f"Save file: {new_file_name}")
        
        # โหลดข้อมูลจากไฟล์ที่สร้างมาใหม่
        df_new = pd.read_excel(new_file_name, sheet_name='Sheet1')
        
        # บันทึกเป็นไฟล์ txt โดยกรองออกคอลัมน์ 'Mobile number'
        txt_file_name = os.path.join(sub2_folder, os.path.basename(new_file_name).replace(".xlsx", ".txt")) #เปลี่ยนตำแหน่งไฟล์ตรงนี้
        df_new_filtered = df_new.drop(columns=['Mobile number'], errors='ignore')
        
        # เขียนข้อมูลลงในไฟล์ .txt
        with open(txt_file_name, 'w',encoding='utf-8') as file:
            # เขียน header
            header_line = '|'.join(df_new_filtered.columns)
            file.write(header_line + '\n')
            
            # เติมค่าว่างให้เป็น string ที่ไม่ใช่ NaN
            df_new_filtered = df_new_filtered.fillna('')
            
            # เขียนข้อมูลแถวต่อแถว
            for index, row in df_new_filtered.iterrows():
                line = '|'.join(map(str, row.values))
                file.write(line + '\n')
        
        print(f"Save file txt: {txt_file_name}")

def create_floder(date):
    real_path = r"\172.16.103.200\CSS_JobReport\Report Outbound Campaign Voice AI\Patch Data"

    current_date = date.strftime("%d/%m/%Y")
    date_folder_path = os.path.join(real_path, current_date)

    if not os.path.exists(date_folder_path):
        os.makedirs(date_folder_path)
        print(f"Create folder: {date_folder_path}")
    else:
        print(f"{date_folder_path} already exists.")
    
    folder_main1 = "Patch data"
    folder_main2 = "Text patch data"
    sub1_folder = os.path.join(date_folder_path, folder_main1)

    if not os.path.exists(sub1_folder):
        os.makedirs(sub1_folder)
        print(f"Create folder: {sub1_folder}")
    else:
        print(f"{sub1_folder} already exists.")
    
    sub2_folder = os.path.join(date_folder_path, folder_main2)
    if not os.path.exists(sub2_folder):
        os.makedirs(sub2_folder)
        print(f"Create folder: {sub2_folder}")
    else:
        print(f"{sub2_folder} already exists.")
    

    

def run_patch():
    patch_data()
    word_to_change = 'System Conect' #เปลี่ยนคำใน call outcome ใน ' ' นี้นะพี่
    change_outcome(word_to_change)
    split_data()

if __name__ == '__main__':
    # Play_premium()
    patch_data()
    word_to_change = 'System Conect' #เปลี่ยนคำใน call outcome ใน ' ' นี้นะพี่
    change_outcome(word_to_change)
    split_data()